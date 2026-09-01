#!/usr/bin/env python3
"""Extract the GO:0006814 candidate IDs from Cao et al. (2023) Table S6."""

from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "references/population_2023/extracted/Supplemental_Tables.xlsx"
OUT = ROOT / "results/population_selected_sodium_transport_genes.tsv"


def main() -> None:
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook["Table S6"]
    header = None
    target = None
    for row in sheet.iter_rows(values_only=True):
        if row and row[0] == "ID":
            header = list(row)
        elif row and row[0] == "GO:0006814":
            if header is None:
                raise RuntimeError("Table header not found before GO:0006814 row")
            target = dict(zip(header, row))
            break
    if target is None:
        raise RuntimeError("GO:0006814 not found in Table S6")

    rows = []
    for gene_id in str(target["geneID"]).split("/"):
        match = re.search(r"scaffold_(\d+)", gene_id)
        rows.append(
            {
                "go_id": target["ID"],
                "description": target["Description"],
                "input_number": target["Input number"],
                "background_number": target["Background number"],
                "pvalue": target["pvalue"],
                "maker_gene_id": gene_id,
                "eastern_original_scaffold": f"scaffold_{match.group(1)}" if match else "unparsed",
            }
        )
    OUT.parent.mkdir(exist_ok=True)
    with OUT.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=rows[0].keys(), delimiter="\t")
        writer.writeheader()
        writer.writerows(rows)
    print(f"Extracted {len(rows)} GO:0006814 gene models from Table S6")


if __name__ == "__main__":
    main()
